"""Excel document (.xlsx) validation with AI write-back"""
import re
import logging


def _normalise_issue(item, rule=None):
    if isinstance(item, dict) and 'rule_name' in item:
        return item
    return {
        'rule_name': rule.get('title', 'Unknown') if rule else 'Unknown',
        'rule_type': rule.get('rule_type', 'Unknown') if rule else 'Unknown',
        'description': str(item),
        'location': 'Workbook-wide',
        'priority': rule.get('priority', 999) if rule else 999
    }


def _normalise_fix(item, rule=None):
    if isinstance(item, dict) and 'rule_name' in item:
        return item
    return {
        'rule_name': rule.get('title', 'Unknown') if rule else 'Unknown',
        'rule_type': rule.get('rule_type', 'Unknown') if rule else 'Unknown',
        'found_value': 'Non-compliant value',
        'fixed_value': str(item),
        'location': 'Workbook-wide'
    }


def validate_excel_document(file_stream, rules):
    """Validate Excel document against rules"""
    from openpyxl import load_workbook
    logging.info("Loading Excel document...")
    wb = load_workbook(file_stream)
    logging.info(f"Excel document loaded. Sheets: {len(wb.sheetnames)}")

    issues = []
    fixes_applied = []

    excel_rules = [r for r in rules if r['doc_type'] in ['Excel', 'Both', 'All']]
    ai_rules = [r for r in excel_rules if r.get('use_ai', False)]
    hard_coded_rules = [r for r in excel_rules if not r.get('use_ai', False)]

    logging.info(f"Hard-coded rules: {len(hard_coded_rules)} (AI rules skipped for Excel)")

    # Hard-coded rules only — AI validation is skipped for Excel
    # (AI is designed for prose; spreadsheet cell text produces too many false positives)
    for rule in hard_coded_rules:
        result = None
        if rule['rule_type'] == 'Font':
            result = _check_fonts(wb, rule)
        elif rule['rule_type'] in ['Language', 'Grammar', 'Punctuation']:
            result = _check_text(wb, rule)

        if result:
            for item in result.get('issues', []):
                issues.append(_normalise_issue(item, rule))
            for item in result.get('fixes', []):
                fixes_applied.append(_normalise_fix(item, rule))

    logging.info(f"Excel validation complete. Issues: {len(issues)}, Fixes: {len(fixes_applied)}")
    return {'document': wb, 'issues': issues, 'fixes_applied': fixes_applied}


def _check_fonts(wb, rule):
    """Check and fix font issues in Excel workbook"""
    from openpyxl.styles import Font
    from copy import copy
    issues = []
    fixes = []
    expected_font = rule['expected_value']

    if rule['check_value'] == 'AllTextFont':
        issue_count = 0
        fix_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        current_font = cell.font.name if cell.font else None
                        if current_font is None or current_font != expected_font:
                            issue_count += 1
                            if rule['auto_fix']:
                                old_font = copy(cell.font)
                                cell.font = Font(
                                    name=expected_font,
                                    size=old_font.size,
                                    bold=old_font.bold,
                                    italic=old_font.italic,
                                    underline=old_font.underline,
                                    color=old_font.color
                                )
                                fix_count += 1

        if issue_count > 0 and not rule['auto_fix']:
            issues.append({
                'rule_name': rule.get('title', 'All Text Font'),
                'rule_type': rule['rule_type'],
                'description': f"Found {issue_count} cells with incorrect font (not {expected_font})",
                'location': 'Workbook-wide',
                'priority': rule.get('priority', 999)
            })
        if fix_count > 0:
            fixes.append({
                'rule_name': rule.get('title', 'All Text Font'),
                'rule_type': rule['rule_type'],
                'found_value': f'{issue_count} cells with wrong font',
                'fixed_value': expected_font,
                'location': f'Workbook-wide ({fix_count} cells)'
            })

    return {'issues': issues, 'fixes': fixes}


def _check_text(wb, rule):
    """Check and fix text issues in Excel (spelling, contractions, symbols, numbers)"""
    issues = []
    fixes = []
    check_value = rule['check_value']
    issue_count = 0
    fix_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                text = cell.value

                if check_value.startswith('BritishSpelling_'):
                    american_word = check_value.replace('BritishSpelling_', '')
                    british_word = rule['expected_value']
                    pattern = r'\b' + re.escape(american_word) + r'\b'
                    matches = re.findall(pattern, text, re.IGNORECASE)
                    if matches:
                        issue_count += len(matches)
                        if rule['auto_fix']:
                            def replace_preserve_case(match, replacement=british_word):
                                word = match.group(0)
                                if word.isupper():
                                    return replacement.upper()
                                elif word[0].isupper():
                                    return replacement.capitalize()
                                return replacement
                            cell.value = re.sub(pattern, replace_preserve_case, text, flags=re.IGNORECASE)
                            fix_count += len(matches)

                elif check_value.startswith('NoContraction_'):
                    contraction = check_value.replace('NoContraction_', '')
                    expanded = rule['expected_value']
                    if contraction in text:
                        count = text.count(contraction)
                        issue_count += count
                        if rule['auto_fix']:
                            cell.value = text.replace(contraction, expanded)
                            fix_count += count

                elif check_value == 'NoAmpersand':
                    if '&' in text:
                        count = text.count('&')
                        issue_count += count
                        if rule['auto_fix']:
                            cell.value = text.replace('&', 'and')
                            fix_count += count

                elif check_value == 'PercentSymbol':
                    percent_matches = re.findall(r'\d+%', text)
                    if percent_matches:
                        issue_count += len(percent_matches)
                        if rule['auto_fix']:
                            cell.value = re.sub(r'(\d+)%', r'\1 percent', text)
                            fix_count += len(percent_matches)

                elif check_value == 'NoApostrophePlurals':
                    apos_matches = re.findall(r"\b[A-Z]{2,}'s\b", text)
                    if apos_matches:
                        issue_count += len(apos_matches)

                elif check_value == 'NumberCommas':
                    num_matches = re.findall(r'\b\d{4,}\b', text)
                    num_matches = [m for m in num_matches if not (1900 <= int(m) <= 2099)]
                    if num_matches:
                        issue_count += len(num_matches)
                        if rule['auto_fix']:
                            for match in num_matches:
                                formatted = '{:,}'.format(int(match))
                                cell.value = cell.value.replace(match, formatted)
                                fix_count += 1

                elif check_value == 'Word_toward':
                    toward_matches = re.findall(r'\btowards\b', text, re.IGNORECASE)
                    if toward_matches:
                        issue_count += len(toward_matches)
                        if rule['auto_fix']:
                            cell.value = re.sub(r'\btowards\b', 'toward', text, flags=re.IGNORECASE)
                            fix_count += len(toward_matches)

                elif check_value == 'AvoidEtc':
                    etc_matches = re.findall(r'\betc\.?\b', text, re.IGNORECASE)
                    if etc_matches:
                        issue_count += len(etc_matches)

    label = rule.get('title', check_value)
    if issue_count > 0:
        issues.append(f"Found {issue_count} instances of '{label}' violations")
    if fix_count > 0:
        fixes.append(f"Fixed {fix_count} instances for '{label}'")

    return {'issues': issues, 'fixes': fixes}
